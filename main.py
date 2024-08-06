import math
import re
import subprocess
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import openpyxl
from openpyxl.worksheet.hyperlink import Hyperlink


# Function to get cell value from Excel using AppleScript
def get_cell_value(cell, max_retries=3, delay=2):
    script = f'''
    tell application "Microsoft Excel"
        try
            set theValue to value of range "{cell}" of worksheet "Tisztítandó adatok" of workbook "/Users/leventebotos/Downloads/test.xlsx"
            return theValue
        on error errMsg number errNum
            return "Error: " & errNum & " - " & errMsg
        end try
    end tell
    '''
    for attempt in range(max_retries):
        try:
            result = subprocess.run(['osascript', '-e', script], capture_output=True, text=True, check=True)
            output = result.stdout.strip()
            if output.startswith("Error"):
                print(f"Attempt {attempt + 1} failed: {output}")
            else:
                return output
        except subprocess.CalledProcessError as e:
            print(f"Attempt {attempt + 1} failed: {e}")
        time.sleep(delay)
    print("All attempts failed. Returning None.")
    return None

# Set up the WebDriver
driver = webdriver.Chrome()

start_row = 11


# Load the Excel workbook to store results
output_workbook = openpyxl.load_workbook('/Users/leventebotos/Downloads/test.xlsx')
output_sheet = output_workbook['Minta']  # Ensure you refer to the correct sheet

# Loop to automate the task
for i in range(3, 141):
    cell = f'B{i}'  # Adjust this according to your needs
    contact = f'C{i}'  # Adjust this according to your needs
    contact_value = get_cell_value(contact)
    input_value = get_cell_value(cell)
    print(input_value)
    print(contact_value)
    
    # Open the website and perform the task
    driver.get('https://www.nemzeticegtar.hu')
    input_field = driver.find_element(By.NAME, 'search')  # Adjust the selector as needed
    input_field.clear()
    input_field.send_keys(input_value)
    input_field.send_keys(Keys.RETURN)
    
    # Wait for the result to load (adjust the time as needed)
    time.sleep(5)

   
    if driver.current_url == 'https://www.nemzeticegtar.hu/nemzeticegtar/lista':
        print("there")
        elements = driver.find_elements(By.CLASS_NAME, "talalat")
        if elements:
            print(elements[0])
            e = elements[0].find_element(By.TAG_NAME, "a")
            e.click()
            time.sleep(5)

       

    # element = driver.find_element(By.XPATH, '//div[class="talalat mb-lg-4 mb-3"]')
    # if element:
    #     elementt= element.find_element(By.TAG_NAME, 'a')
    #     elementt.click()

    # top = driver.find_elements(By.CLASS_NAME, 'talalat mb-lg-4 mb-3')
    # print('found')
    # if top:
    #     topp = top[0].find_element(By.TAG_NAME, "a")
    #     print('found')
    #     if topp:
    #         topp.click()
    #         print("clicked Top")
    #         time.sleep(5)
        

      
    
    # Initialize a list to store the row data
    row_data = []
    
    try:
       
    #     talalat = driver.find_element(By.XPATH, '//div[@class="talalat"]/div')
    
    # # Check if the element is found and clickable, then click it
    #     if talalat:
    #         talalat.click()
    #         time.sleep(5)

        

        # Extract the current URL
        current_url = driver.current_url
        
        # Extract the company name
        company_name = driver.find_element(By.XPATH, '//div[@class="col-12 bg-light mb-0 pt-4 pt-lg-0"]/h1').text
        row_data.extend(['', '', company_name, 'HU'])

        tax = driver.find_element(By.XPATH, '//h6[text()="Adószám"]/following-sibling::p').text
        row_data.append(tax)
        
        row_data.extend(["", "", "", "Hungary"])
        
        address = driver.find_element(By.XPATH, '//h6[text()="székhely"]/following-sibling::p').text
        address = address.rstrip('.')
        
        # Split the string by comma
        zip_city, street_and_number = address.split(', ')
        zip_code, city = zip_city.split(' ', 1)
        street_and_number_parts = street_and_number.split(' ')
        street = ' '.join(street_and_number_parts[:-1])
        number = street_and_number_parts[-1]
        
        row_data.extend([zip_code, city, street, number, "", "", "", ""])
        
        main = driver.find_element(By.XPATH, '//h6[text()="Főtevékenység"]/following-sibling::p').text
        row_data.append(main[:4])
        
        # Add the current URL as a clickable hyperlink
     
        
        # Extract net revenue
        try:
            # net_revenue = driver.find_element(By.XPATH, '//div[contains(@class, "rounded-information-item")]//span').text
            # row_data.append(net_revenue)
            h5_element = driver.find_element(By.XPATH, '//h5[contains(text(), "nettó árbevétel")]')
    
    # Find the parent `div` with class "rounded-information-item"
            parent_div = h5_element.find_element(By.XPATH, './following-sibling::div[contains(@class, "rounded-information-item")]')
    
    # Find the nested `span` inside the parent `div`
            net_revenue_span = parent_div.find_element(By.XPATH, './/div[@class="inner"]//span')
    
    # Extract the text from the `span` element
            net_revenue = net_revenue_span.text

            net_revenue = net_revenue[:-8].strip()

        

    # Append the net revenue to the row_data list
            row_data.append(net_revenue)
        except Exception as e:
            print(f"Error retrieving net revenue for cell B{i}: {e}")
            row_data.append('Error')
           


        
        # Extract min and max people, then calculate the average
        try:
            h5_element2 = driver.find_element(By.XPATH, '//h5[contains(text(), "létszám")]')
    
    # Find the parent `div` with class "rounded-information-item"
            parent_div2 = h5_element2.find_element(By.XPATH, './following-sibling::div[contains(@class, "rounded-information-item")]')
    
    # Find the nested `span` inside the parent `div`
            people_span = parent_div2.find_element(By.XPATH, './/div[@class="inner"]//span')
    
    # Extract the text from the `span` element
            people_text = people_span.text
    
         
            print(f"Raw people text: '{people_text}'")  # Print raw text for debugging
            
            # Remove unwanted characters and split range
            people_text = people_text.strip().replace('fő', '').replace(' ', '')
            if "-" in people_text:
                people_range = people_text.split('-')
                min_people = int(people_range[0])
                max_people = int(people_range[1])
                average_people = math.ceil((min_people + max_people) / 2)
                
                row_data.append(average_people)
            else:
                print(people_text)
                raise ValueError("Unexpected format for people range")
        
        except Exception as e:
            print(f"Error retrieving people numbers for cell B{i}: {e}")
            row_data.append('Error')
       
          


        row_data.append("")


        # url_cell = output_sheet.cell(row=i, column=len(row_data) + 1)
        # url_cell.value = "Goooleee"
        # url_cell.hyperlink = Hyperlink("https://google.com", "Google")
        row_data.append(current_url)

        
        # Append contact value
        row_data.append(contact_value)
        
        # Write the data to the "Minta" sheet starting from row 11
        for col, value in enumerate(row_data, start=1):
            output_sheet.cell(row=start_row, column=col).value = value
        
        # Increment the row index for the next set of data
        start_row += 1
    
    except Exception as e:
        print(f"Error retrieving data for cell B{i}: {e}")
        output_sheet.cell(row=start_row, column=1).value = 'Error'
        output_sheet.cell(row=start_row, column=3).value = input_value
        output_sheet.cell(row=start_row, column=23).value = contact_value
        start_row += 1

# Save the output workbook
output_workbook.save('/Users/leventebotos/Downloads/test_output.xlsx')

# Close the driver
driver.quit()
